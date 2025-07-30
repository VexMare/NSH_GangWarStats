#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
帮会联赛数据处理程序 - 高级版本
功能：读取CSV文件，按团长和职业排序，生成Excel文件，包含数据统计
"""

import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import os
from datetime import datetime

class GuildLeagueProcessorAdvanced:
    def __init__(self, csv_file_path):
        self.csv_file_path = csv_file_path
        self.data = []
        self.guild1_data = []
        self.guild2_data = []
        
    def read_csv_data(self):
        """读取CSV文件数据"""
        try:
            with open(self.csv_file_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                self.data = list(reader)
            
            # 找到空行分隔符
            separator_line = None
            for i, row in enumerate(self.data):
                if not row or all(cell.strip() == '' for cell in row):
                    separator_line = i
                    break
            
            if separator_line is None:
                print("警告：未找到空行分隔符，使用默认第92行")
                separator_line = 91
            
            # 分离两个帮会的数据
            self.guild1_data = self.data[1:separator_line]  # 跳过标题行
            self.guild2_data = self.data[separator_line + 1:]  # 跳过空行
            
            print(f"成功读取数据：")
            print(f"帮会1数据行数：{len(self.guild1_data)}")
            print(f"帮会2数据行数：{len(self.guild2_data)}")
            
        except Exception as e:
            print(f"读取CSV文件时出错：{e}")
            return False
        return True
    
    def create_dataframe(self, data, guild_name):
        """创建DataFrame"""
        columns = ['帮会名', '玩家', '等级', '职业', '所在团长', '击败', '助攻', '战备资源', 
                  '对玩家伤害', '对建筑伤害', '治疗值', '承受伤害', '重伤', '青灯焚骨', '化羽', '控制']
        
        df = pd.DataFrame(data, columns=columns)
        df['帮会名'] = guild_name
        
        # 转换数值列
        numeric_columns = ['等级', '击败', '助攻', '战备资源', '对玩家伤害', '对建筑伤害', 
                          '治疗值', '承受伤害', '重伤', '青灯焚骨', '化羽', '控制']
        
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        return df
    
    def sort_by_leader(self, df):
        """按团长排序，添加分割线"""
        sorted_df = df.sort_values(['所在团长', '对玩家伤害'], ascending=[True, False])
        
        # 创建带分割线的数据
        result_list = []
        current_leader = None
        
        for _, row in sorted_df.iterrows():
            if row['所在团长'] != current_leader:
                # 添加分割线
                if current_leader is not None:  # 不是第一个团长
                    # 添加空行
                    empty_row = pd.DataFrame([{
                        '帮会名': '',
                        '玩家': '',
                        '等级': '',
                        '职业': '',
                        '所在团长': '',
                        '击败': '',
                        '助攻': '',
                        '战备资源': '',
                        '对玩家伤害': '',
                        '对建筑伤害': '',
                        '治疗值': '',
                        '承受伤害': '',
                        '重伤': '',
                        '青灯焚骨': '',
                        '化羽': '',
                        '控制': ''
                    }])
                    result_list.append(empty_row)
                
                # 添加标题行
                title_row = pd.DataFrame([{
                    '帮会名': '帮会名',
                    '玩家': '玩家',
                    '等级': '等级',
                    '职业': '职业',
                    '所在团长': '所在团长',
                    '击败': '击败',
                    '助攻': '助攻',
                    '战备资源': '战备资源',
                    '对玩家伤害': '对玩家伤害',
                    '对建筑伤害': '对建筑伤害',
                    '治疗值': '治疗值',
                    '承受伤害': '承受伤害',
                    '重伤': '重伤',
                    '青灯焚骨': '青灯焚骨',
                    '化羽': '化羽',
                    '控制': '控制'
                }])
                result_list.append(title_row)
                current_leader = row['所在团长']
            
            # 添加数据行
            result_list.append(pd.DataFrame([row]))
        
        return pd.concat(result_list, ignore_index=True)
    
    def sort_by_profession(self, df):
        """按职业排序，添加分割线"""
        # 根据不同职业使用不同的排序指标
        def get_sort_key(row):
            profession = row['职业']
            if profession == '素问':
                return row['治疗值']
            elif profession == '九灵':
                return row['青灯焚骨']
            else:
                return row['对玩家伤害']
        
        # 添加排序键列
        df_with_sort_key = df.copy()
        df_with_sort_key['sort_key'] = df_with_sort_key.apply(get_sort_key, axis=1)
        
        # 按职业和排序键排序
        sorted_df = df_with_sort_key.sort_values(['职业', 'sort_key'], ascending=[True, False])
        
        # 创建带分割线的数据
        result_list = []
        current_profession = None
        
        for _, row in sorted_df.iterrows():
            if row['职业'] != current_profession:
                # 添加分割线
                if current_profession is not None:  # 不是第一个职业
                    # 添加空行
                    empty_row = pd.DataFrame([{
                        '帮会名': '',
                        '玩家': '',
                        '等级': '',
                        '职业': '',
                        '所在团长': '',
                        '击败': '',
                        '助攻': '',
                        '战备资源': '',
                        '对玩家伤害': '',
                        '对建筑伤害': '',
                        '治疗值': '',
                        '承受伤害': '',
                        '重伤': '',
                        '青灯焚骨': '',
                        '化羽': '',
                        '控制': ''
                    }])
                    result_list.append(empty_row)
                
                # 添加标题行
                title_row = pd.DataFrame([{
                    '帮会名': '帮会名',
                    '玩家': '玩家',
                    '等级': '等级',
                    '职业': '职业',
                    '所在团长': '所在团长',
                    '击败': '击败',
                    '助攻': '助攻',
                    '战备资源': '战备资源',
                    '对玩家伤害': '对玩家伤害',
                    '对建筑伤害': '对建筑伤害',
                    '治疗值': '治疗值',
                    '承受伤害': '承受伤害',
                    '重伤': '重伤',
                    '青灯焚骨': '青灯焚骨',
                    '化羽': '化羽',
                    '控制': '控制'
                }])
                result_list.append(title_row)
                current_profession = row['职业']
            
            # 添加数据行（移除排序键列）
            data_row = row.drop('sort_key')
            result_list.append(pd.DataFrame([data_row]))
        
        return pd.concat(result_list, ignore_index=True)
    
    def create_statistics(self, df, guild_name):
        """创建统计数据"""
        stats = {
            '帮会名': guild_name,
            '总人数': len(df),
            '总击败数': df['击败'].sum(),
            '总助攻数': df['助攻'].sum(),
            '总战备资源': df['战备资源'].sum(),
            '总对玩家伤害': df['对玩家伤害'].sum(),
            '总对建筑伤害': df['对建筑伤害'].sum(),
            '总治疗值': df['治疗值'].sum(),
            '总承受伤害': df['承受伤害'].sum(),
            '总重伤数': df['重伤'].sum(),
            '总青灯焚骨': df['青灯焚骨'].sum(),
            '总化羽数': df['化羽'].sum(),
            '总控制数': df['控制'].sum(),
        }
        
        return stats
    
    def create_profession_statistics(self, df):
        """创建职业统计数据，按职业分别显示"""
        profession_stats_list = []
        for profession in sorted(df['职业'].unique()):
            profession_data = df[df['职业'] == profession]
            
            # 对职业数据进行排序（按治疗值排序素问，按青灯焚骨排序九灵，其他按对玩家伤害排序）
            def get_sort_key(row):
                if profession == '素问':
                    return row['治疗值']
                elif profession == '九灵':
                    return row['青灯焚骨']
                else:
                    return row['对玩家伤害']
            
            profession_data_sorted = profession_data.copy()
            profession_data_sorted['sort_key'] = profession_data_sorted.apply(get_sort_key, axis=1)
            profession_data_sorted = profession_data_sorted.sort_values('sort_key', ascending=False)
            profession_data_sorted = profession_data_sorted.drop('sort_key', axis=1)
            
            # 添加职业标题行
            title_row = pd.DataFrame([{col: '' for col in df.columns}])
            title_row.iloc[0, df.columns.get_loc('职业')] = f'=== {profession} ==='
            # 添加该职业的统计数据
            stats_row = pd.DataFrame([{col: '' for col in df.columns}])
            stats_row.iloc[0, df.columns.get_loc('职业')] = f'{profession}统计'
            stats_row.iloc[0, df.columns.get_loc('玩家')] = f'人数: {len(profession_data)}'
            stats_row.iloc[0, df.columns.get_loc('等级')] = f'平均: {profession_data["等级"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('击败')] = f'总计: {profession_data["击败"].sum()}, 平均: {profession_data["击败"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('助攻')] = f'总计: {profession_data["助攻"].sum()}, 平均: {profession_data["助攻"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('战备资源')] = f'总计: {profession_data["战备资源"].sum()}'
            stats_row.iloc[0, df.columns.get_loc('对玩家伤害')] = f'总计: {profession_data["对玩家伤害"].sum():,.0f}, 平均: {profession_data["对玩家伤害"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('对建筑伤害')] = f'总计: {profession_data["对建筑伤害"].sum():,.0f}, 平均: {profession_data["对建筑伤害"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('治疗值')] = f'总计: {profession_data["治疗值"].sum():,.0f}, 平均: {profession_data["治疗值"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('承受伤害')] = f'总计: {profession_data["承受伤害"].sum():,.0f}, 平均: {profession_data["承受伤害"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('重伤')] = f'总计: {profession_data["重伤"].sum()}, 平均: {profession_data["重伤"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('青灯焚骨')] = f'总计: {profession_data["青灯焚骨"].sum()}, 平均: {profession_data["青灯焚骨"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('化羽')] = f'总计: {profession_data["化羽"].sum()}, 平均: {profession_data["化羽"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('控制')] = f'总计: {profession_data["控制"].sum()}, 平均: {profession_data["控制"].mean():.1f}'
            # 添加列标题行
            header_row = pd.DataFrame([{col: col for col in df.columns}])
            # 详细数据 - 显示所有玩家的完整数据（已排序）
            detail = profession_data_sorted.copy()
            profession_stats_list.extend([title_row, stats_row, header_row, detail])
            # 添加空行作为间隔
            if profession != sorted(df['职业'].unique())[-1]:
                empty_row = pd.DataFrame([{col: '' for col in df.columns}])
                profession_stats_list.append(empty_row)
        return pd.concat(profession_stats_list, ignore_index=True)

    def create_leader_statistics(self, df):
        """创建团长统计数据，按团长分别显示"""
        leader_stats_list = []
        for leader in sorted(df['所在团长'].unique()):
            leader_data = df[df['所在团长'] == leader]
            
            # 对团长数据进行排序（按对玩家伤害排序）
            leader_data_sorted = leader_data.sort_values('对玩家伤害', ascending=False)
            
            # 添加团长标题行
            title_row = pd.DataFrame([{col: '' for col in df.columns}])
            title_row.iloc[0, df.columns.get_loc('所在团长')] = f'=== {leader} ==='
            # 添加该团长的统计数据
            stats_row = pd.DataFrame([{col: '' for col in df.columns}])
            stats_row.iloc[0, df.columns.get_loc('所在团长')] = f'{leader}统计'
            stats_row.iloc[0, df.columns.get_loc('击败')] = f'总计: {leader_data["击败"].sum()}, 平均: {leader_data["击败"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('助攻')] = f'总计: {leader_data["助攻"].sum()}, 平均: {leader_data["助攻"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('战备资源')] = f'总计: {leader_data["战备资源"].sum()}'
            stats_row.iloc[0, df.columns.get_loc('对玩家伤害')] = f'总计: {leader_data["对玩家伤害"].sum():,.0f}, 平均: {leader_data["对玩家伤害"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('对建筑伤害')] = f'总计: {leader_data["对建筑伤害"].sum():,.0f}, 平均: {leader_data["对建筑伤害"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('治疗值')] = f'总计: {leader_data["治疗值"].sum():,.0f}, 平均: {leader_data["治疗值"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('承受伤害')] = f'总计: {leader_data["承受伤害"].sum():,.0f}, 平均: {leader_data["承受伤害"].mean():,.0f}'
            stats_row.iloc[0, df.columns.get_loc('重伤')] = f'总计: {leader_data["重伤"].sum()}, 平均: {leader_data["重伤"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('青灯焚骨')] = f'总计: {leader_data["青灯焚骨"].sum()}, 平均: {leader_data["青灯焚骨"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('化羽')] = f'总计: {leader_data["化羽"].sum()}, 平均: {leader_data["化羽"].mean():.1f}'
            stats_row.iloc[0, df.columns.get_loc('控制')] = f'总计: {leader_data["控制"].sum()}, 平均: {leader_data["控制"].mean():.1f}'
            # 添加列标题行
            header_row = pd.DataFrame([{col: col for col in df.columns}])
            # 详细数据 - 显示所有玩家的完整数据（已排序）
            detail = leader_data_sorted.copy()
            leader_stats_list.extend([title_row, stats_row, header_row, detail])
            # 添加空行作为间隔
            if leader != sorted(df['所在团长'].unique())[-1]:
                empty_row = pd.DataFrame([{col: '' for col in df.columns}])
                leader_stats_list.append(empty_row)
        return pd.concat(leader_stats_list, ignore_index=True)
    
    def create_excel_file(self, output_file):
        """创建Excel文件"""
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建数据框
        guild1_df = self.create_dataframe(self.guild1_data, "山有扶苏")
        guild2_df = self.create_dataframe(self.guild2_data, "璃月")
        
        # 合并数据用于综合职业排序
        combined_df = pd.concat([guild1_df, guild2_df], ignore_index=True)
        
        # 创建统计数据
        stats1 = self.create_statistics(guild1_df, "山有扶苏")
        stats2 = self.create_statistics(guild2_df, "璃月")
        
        # 创建职业和团长统计
        prof_stats1 = self.create_profession_statistics(guild1_df)
        prof_stats2 = self.create_profession_statistics(guild2_df)
        leader_stats1 = self.create_leader_statistics(guild1_df)
        leader_stats2 = self.create_leader_statistics(guild2_df)
        
        # 首先创建广告页面
        ws_ad = wb.create_sheet(title="关于程序", index=0)
        self.create_advertisement_page(ws_ad)
        
        # 创建工作表
        sheets = [
            ("本帮团长排序", self.sort_by_leader(guild1_df)),
            ("本帮职业排序", self.sort_by_profession(guild1_df)),
            ("敌帮团长排序", self.sort_by_leader(guild2_df)),
            ("敌帮职业排序", self.sort_by_profession(guild2_df)),
            ("综合职业排序", self.sort_by_profession(combined_df)),
            ("本帮职业统计", prof_stats1),
            ("本帮团长统计", leader_stats1),
            ("敌帮职业统计", prof_stats2),
            ("敌帮团长统计", leader_stats2),
            ("帮会对比", pd.DataFrame([stats1, stats2]))
        ]
        
        for sheet_name, df in sheets:
            ws = wb.create_sheet(title=sheet_name)
            self.format_worksheet(ws, df, sheet_name)
        
        # 保存文件
        wb.save(output_file)
        print(f"Excel文件已保存：{output_file}")
    
    def create_advertisement_page(self, ws):
        """创建广告页面"""
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
        
        # 标题样式
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 副标题样式
        subtitle_font = Font(bold=True, size=14, color="FFFFFF")
        subtitle_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subtitle_alignment = Alignment(horizontal="center", vertical="center")
        
        # 正文样式
        content_font = Font(size=11)
        content_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # 链接样式
        link_font = Font(size=11, color="0000FF", underline="single")
        
        # 边框样式
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 添加内容
        content_data = [
            # 标题
            ("", "帮会联赛数据处理程序 - 高级版", ""),
            ("", "", ""),
            
            # 程序信息
            ("", "程序信息", ""),
            ("版本", "高级版 V2.0", ""),
            ("作者", "VexMare（恶小梦）", ""),
            ("更新时间", "2025.7.30", ""),
            ("", "", ""),
            
            # 联系方式
            ("", "联系方式", ""),
            ("GitHub", "https://github.com/VexMare/NSH_GangWarStats", ""),
            ("BiliBili", "https://space.bilibili.com/365374856?spm_id_from=333.1007.0.0", ""),
            ("邮箱", "chixiaotao@foxmail.com", ""),
            ("", "", ""),
            
            # 功能特色
            ("", "功能特色", ""),
            ("", "• 智能数据排序：按团长、职业等多种方式排序", ""),
            ("", "• 颜色可视化：不同数据用不同颜色突出显示", ""),
            ("", "• 详细统计：包含总计、平均值等统计信息", ""),
            ("", "• 多工作表：10个不同视角的数据分析", ""),
            ("", "• 职业特定排序：素问按治疗值，九灵按青灯焚骨", ""),
            ("", "", ""),
            
            # 使用说明
            ("", "使用说明", ""),
            ("", "1. 准备CSV文件，确保格式正确（UTF-8编码）", ""),
            ("", "2. 运行程序，选择CSV文件", ""),
            ("", "3. 程序自动生成Excel文件，包含10个工作表", ""),
            ("", "4. 查看不同角度的数据分析结果", ""),
            ("", "", ""),
            
            # 版权声明
            ("", "版权声明", ""),
            ("", "本程序采用 MIT 许可证", ""),
            ("", "", ""),
            ("", "MIT License", ""),
            ("", "Copyright (c) 2024 VexMare", ""),
            ("", "", ""),
            ("", "Permission is hereby granted, free of charge, to any person obtaining a copy", ""),
            ("", "of this software and associated documentation files (the \"Software\"), to deal", ""),
            ("", "in the Software without restriction, including without limitation the rights", ""),
            ("", "to use, copy, modify, merge, publish, distribute, sublicense, and/or sell", ""),
            ("", "copies of the Software, and to permit persons to whom the Software is", ""),
            ("", "furnished to do so, subject to the following conditions:", ""),
            ("", "", ""),
            ("", "The above copyright notice and this permission notice shall be included in all", ""),
            ("", "copies or substantial portions of the Software.", ""),
            ("", "", ""),
            ("", "THE SOFTWARE IS PROVIDED \"AS IS\", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR", ""),
            ("", "IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,", ""),
            ("", "FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE", ""),
            ("", "AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER", ""),
            ("", "LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,", ""),
            ("", "OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE", ""),
            ("", "SOFTWARE.", ""),
            ("", "", ""),
            
            # 重要声明
            ("", "重要声明", ""),
            ("", "⚠️  禁止商用：本程序仅供学习和个人使用，禁止用于商业用途", ""),
            ("", "⚠️  免责声明：使用本程序产生的任何后果由用户自行承担", ""),
            ("", "⚠️  技术支持：如有问题，请通过上述联系方式联系作者", ""),
            ("", "", ""),
            
            # 未来计划
            ("", "未来计划", ""),
            ("", "我们即将推出收费版网站服务，提供更便捷的在线数据处理功能。", ""),
            ("", "如果您有相关开发经验，欢迎加入我们的团队！", ""),
            ("", "请将您的GitHub链接发送到邮箱：chixiaotao@foxmail.com", ""),
        ]
        
        # 写入内容并设置样式
        for row_idx, (col_a, col_b, col_c) in enumerate(content_data, 1):
            # 写入数据
            ws.cell(row=row_idx, column=1, value=col_a)
            ws.cell(row=row_idx, column=2, value=col_b)
            ws.cell(row=row_idx, column=3, value=col_c)
            
            # 设置边框
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = border
            
            # 根据内容设置样式
            if "帮会联赛数据处理程序" in str(col_b):
                # 主标题
                ws.cell(row=row_idx, column=2).font = title_font
                ws.cell(row=row_idx, column=2).fill = title_fill
                ws.cell(row=row_idx, column=2).alignment = title_alignment
                # 合并单元格
                ws.merge_cells(f'A{row_idx}:C{row_idx}')
            elif any(keyword in str(col_b) for keyword in ["程序信息", "联系方式", "功能特色", "使用说明", "版权声明", "重要声明", "未来计划"]):
                # 副标题
                ws.cell(row=row_idx, column=2).font = subtitle_font
                ws.cell(row=row_idx, column=2).fill = subtitle_fill
                ws.cell(row=row_idx, column=2).alignment = subtitle_alignment
                # 合并单元格
                ws.merge_cells(f'A{row_idx}:C{row_idx}')
            elif any(keyword in str(col_b) for keyword in ["https://", "chixiaotao@foxmail.com"]):
                # 链接
                ws.cell(row=row_idx, column=2).font = link_font
                ws.cell(row=row_idx, column=2).alignment = content_alignment
            elif col_b and col_b != "":
                # 正文内容
                ws.cell(row=row_idx, column=2).font = content_font
                ws.cell(row=row_idx, column=2).alignment = content_alignment
        
        # 设置行高
        for row in range(1, len(content_data) + 1):
            if "帮会联赛数据处理程序" in str(ws.cell(row=row, column=2).value):
                ws.row_dimensions[row].height = 30
            elif any(keyword in str(ws.cell(row=row, column=2).value) for keyword in ["程序信息", "联系方式", "功能特色", "使用说明", "版权声明", "重要声明", "未来计划"]):
                ws.row_dimensions[row].height = 25
            else:
                ws.row_dimensions[row].height = 18
    
    def format_worksheet(self, ws, df, sheet_name):
        """格式化工作表"""
        # 设置列宽
        column_widths = {
            'A': 12, 'B': 14, 'C': 8, 'D': 10, 'E': 14, 'F': 12, 'G': 12, 'H': 14, 'I': 14, 'J': 14, 'K': 14, 'L': 14, 'M': 14, 'N': 14, 'O': 14, 'P': 14
        }
        # 针对统计表，设置指定列宽为24
        if any(x in sheet_name for x in ["团长统计", "职业统计"]):
            for col in ['F', 'G', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                ws.column_dimensions[col].width = 27
        else:
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
        
        # 安全处理值，避免公式问题
        def safe_value(value):
            """安全处理值，避免公式问题"""
            if isinstance(value, str):
                # 如果字符串以等号开头，添加单引号前缀
                if value.startswith('='):
                    return f"'{value}"
                # 如果字符串包含特殊字符，用引号包围
                elif any(char in value for char in ['+', '-', '*', '/', '(', ')', '=']):
                    return f"'{value}"
            return value
        
        # 添加数据，使用安全值处理
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                cell.value = safe_value(value)
        
        # 设置标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据行样式
        data_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = border
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 为统计表添加特殊格式
        if "统计" in sheet_name:
            self.add_statistics_formatting(ws, df)
            # 为统计表的详细数据部分也添加颜色格式化
            self.add_damage_color_gradient(ws, df)
        # 为排序表添加特殊格式
        elif "排序" in sheet_name:
            self.add_sorting_formatting(ws, df)
    
    def add_statistics_formatting(self, ws, df):
        """为统计表添加特殊格式"""
        # 为标题行添加特殊样式
        title_font = Font(bold=True, color="FFFFFF", size=12)
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # 为统计行添加特殊样式
        stats_font = Font(bold=True, color="FFFFFF")
        stats_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # 为列标题行添加特殊样式
        header_font = Font(bold=True, color="000000", size=11)
        header_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        
        for row in range(2, ws.max_row + 1):
            # 检查是否是标题行（包含"==="的行）
            title_cell = None
            if "职业统计" in ws.title:
                title_cell = ws[f'A{row}']  # 职业列
            elif "团长统计" in ws.title:
                title_cell = ws[f'E{row}']  # 团长列
            
            if title_cell and title_cell.value and '===' in str(title_cell.value):
                # 标题行样式
                for cell in ws[row]:
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            elif title_cell and title_cell.value and '统计' in str(title_cell.value):
                # 统计行样式
                for cell in ws[row]:
                    cell.font = stats_font
                    cell.fill = stats_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            elif title_cell and title_cell.value in ['帮会名', '玩家', '职业', '所在团长']:
                # 列标题行样式
                for cell in ws[row]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def add_sorting_formatting(self, ws, df):
        """为排序表添加特殊格式"""
        # 为标题行添加特殊样式
        header_font = Font(bold=True, color="000000", size=11)
        header_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        
        for row in range(2, ws.max_row + 1):
            # 检查是否是标题行（包含列标题的行）
            first_cell = ws[f'A{row}']
            if first_cell.value == '帮会名':
                # 标题行样式
                for cell in ws[row]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 为I列和J列添加颜色渐变（基于职业组内的伤害百分比）
        self.add_damage_color_gradient(ws, df)
    
    def add_damage_color_gradient(self, ws, df):
        """为伤害列添加颜色渐变"""
        from openpyxl.formatting.rule import DataBarRule
        
        # 判断是按职业分组还是按团长分组
        is_profession_sort = "职业排序" in ws.title
        is_leader_sort = "团长排序" in ws.title
        is_profession_stats = "职业统计" in ws.title
        is_leader_stats = "团长统计" in ws.title
        
        if is_profession_sort or is_profession_stats:
            # 按职业分组
            group_column = 'D'  # 职业列
            group_name = "职业"
        elif is_leader_sort or is_leader_stats:
            # 按团长分组
            group_column = 'E'  # 团长列
            group_name = "团长"
        else:
            return  # 不是排序表或统计表，不处理
        
        # 获取所有分组
        groups = []
        current_group = None
        group_ranges = []
        start_row = 2
        
        for row in range(2, ws.max_row + 1):
            group_cell = ws[f'{group_column}{row}']
            if group_cell.value and group_cell.value != current_group:
                if current_group is not None:
                    # 结束上一个分组
                    group_ranges.append((current_group, start_row, row - 1))
                current_group = group_cell.value
                start_row = row
        
        # 添加最后一个分组
        if current_group is not None:
            group_ranges.append((current_group, start_row, ws.max_row))
        
        # 为每个分组内的不同列设置颜色
        for group, start_row, end_row in group_ranges:
            # 找到该分组内的数据行（排除标题行和空行）
            actual_data_rows = []
            damage_values = []
            
            for row in range(start_row, end_row + 1):
                if ws[f'A{row}'].value and ws[f'A{row}'].value != '帮会名':
                    actual_data_rows.append(row)
                    # 获取各种数值
                    i_damage = ws[f'I{row}'].value  # 对玩家伤害
                    j_damage = ws[f'J{row}'].value  # 对建筑伤害
                    k_heal = ws[f'K{row}'].value    # 治疗值
                    l_tank = ws[f'L{row}'].value    # 承受伤害
                    m_death = ws[f'M{row}'].value   # 重伤
                    n_lantern = ws[f'N{row}'].value # 青灯焚骨
                    p_control = ws[f'P{row}'].value # 控制
                    
                    if isinstance(i_damage, (int, float)):
                        damage_values.append((row, i_damage, j_damage, k_heal, l_tank, m_death, n_lantern, p_control))
            
            if len(damage_values) > 1:
                # 计算该分组内的最大值
                max_i_damage = max(damage_values, key=lambda x: x[1])[1]
                max_j_damage = max(damage_values, key=lambda x: x[2])[2]
                max_k_heal = max(damage_values, key=lambda x: x[3])[3]
                max_l_tank = max(damage_values, key=lambda x: x[4])[4]
                max_m_death = max(damage_values, key=lambda x: x[5])[5]
                max_n_lantern = max(damage_values, key=lambda x: x[6])[6]
                max_p_control = max(damage_values, key=lambda x: x[7])[7]
                
                # 为每个数据行设置颜色
                for row, i_damage, j_damage, k_heal, l_tank, m_death, n_lantern, p_control in damage_values:
                    # F列（击败）- 红色
                    f_kill = ws[f'F{row}'].value  # 击败
                    if isinstance(f_kill, (int, float)) and f_kill > 0:
                        max_f_kill = max(damage_values, key=lambda x: ws[f'F{x[0]}'].value if isinstance(ws[f'F{x[0]}'].value, (int, float)) else 0)[0]
                        max_f_kill_value = ws[f'F{max_f_kill}'].value
                        if max_f_kill_value > 0:
                            f_rule = DataBarRule(
                                start_type='num', start_value=0,
                                end_type='num', end_value=max_f_kill_value,
                                color='FF0000'  # 红色
                            )
                            ws.conditional_formatting.add(f'F{row}', f_rule)
                    
                    # G列（助攻）- 绿色
                    g_assist = ws[f'G{row}'].value  # 助攻
                    if isinstance(g_assist, (int, float)) and g_assist > 0:
                        max_g_assist = max(damage_values, key=lambda x: ws[f'G{x[0]}'].value if isinstance(ws[f'G{x[0]}'].value, (int, float)) else 0)[0]
                        max_g_assist_value = ws[f'G{max_g_assist}'].value
                        if max_g_assist_value > 0:
                            g_rule = DataBarRule(
                                start_type='num', start_value=0,
                                end_type='num', end_value=max_g_assist_value,
                                color='00FF00'  # 绿色
                            )
                            ws.conditional_formatting.add(f'G{row}', g_rule)
                    
                    # I列（对玩家伤害）- 红色
                    if max_i_damage > 0 and i_damage > 0:
                        i_rule = DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=max_i_damage,
                            color='FF0000'  # 红色
                        )
                        ws.conditional_formatting.add(f'I{row}', i_rule)
                    
                    # J列（对建筑伤害）- 黄色
                    if max_j_damage > 0 and j_damage > 0:
                        j_rule = DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=max_j_damage,
                            color='FFFF00'  # 黄色
                        )
                        ws.conditional_formatting.add(f'J{row}', j_rule)
                    
                    # K列（治疗值）- 绿色（仅素问职业）
                    if (is_profession_sort or is_profession_stats) and group == '素问' and max_k_heal > 0 and k_heal > 0:
                        k_rule = DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=max_k_heal,
                            color='00FF00'  # 绿色
                        )
                        ws.conditional_formatting.add(f'K{row}', k_rule)
                    elif is_leader_sort or is_leader_stats:
                        # 在团长排序/统计中，检查该玩家的职业是否为素问
                        profession_cell = ws[f'D{row}']  # 职业列
                        if profession_cell.value == '素问' and max_k_heal > 0 and k_heal > 0:
                            k_rule = DataBarRule(
                                start_type='num', start_value=0,
                                end_type='num', end_value=max_k_heal,
                                color='00FF00'  # 绿色
                            )
                            ws.conditional_formatting.add(f'K{row}', k_rule)
                    
                    # L列（承受伤害）- 浅蓝色
                    if max_l_tank > 0 and l_tank > 0:
                        l_rule = DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=max_l_tank,
                            color='87CEEB'  # 浅蓝色
                        )
                        ws.conditional_formatting.add(f'L{row}', l_rule)
                    
                    # M列（重伤）- 紫色偏红
                    if max_m_death > 0 and m_death > 0:
                        m_rule = DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=max_m_death,
                            color='800080'  # 紫色偏红
                        )
                        ws.conditional_formatting.add(f'M{row}', m_rule)
                    
                    # N列（青灯焚骨）- 紫色（仅九灵职业）
                    if (is_profession_sort or is_profession_stats) and group == '九灵' and max_n_lantern > 0 and n_lantern > 0:
                        n_rule = DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=max_n_lantern,
                            color='800080'  # 紫色
                        )
                        ws.conditional_formatting.add(f'N{row}', n_rule)
                    elif is_leader_sort or is_leader_stats:
                        # 在团长排序/统计中，检查该玩家的职业是否为九灵
                        profession_cell = ws[f'D{row}']  # 职业列
                        if profession_cell.value == '九灵' and max_n_lantern > 0 and n_lantern > 0:
                            n_rule = DataBarRule(
                                start_type='num', start_value=0,
                                end_type='num', end_value=max_n_lantern,
                                color='800080'  # 紫色
                            )
                            ws.conditional_formatting.add(f'N{row}', n_rule)
                    
                    # P列（控制）- 深蓝色
                    if max_p_control > 0 and p_control > 0:
                        p_rule = DataBarRule(
                            start_type='num', start_value=0,
                            end_type='num', end_value=max_p_control,
                            color='000080'  # 深蓝色
                        )
                        ws.conditional_formatting.add(f'P{row}', p_rule)
                    
                    # O列（化羽）- 粉色（仅素问职业）
                    o_feather = ws[f'O{row}'].value  # 化羽
                    if isinstance(o_feather, (int, float)) and o_feather > 0:
                        if (is_profession_sort or is_profession_stats) and group == '素问':
                            # 在职业排序/统计中，素问职业的化羽显示粉色
                            max_o_feather = max(damage_values, key=lambda x: ws[f'O{x[0]}'].value if isinstance(ws[f'O{x[0]}'].value, (int, float)) else 0)[0]
                            max_o_feather_value = ws[f'O{max_o_feather}'].value
                            if max_o_feather_value > 0:
                                o_rule = DataBarRule(
                                    start_type='num', start_value=0,
                                    end_type='num', end_value=max_o_feather_value,
                                    color='FFC0CB'  # 粉色
                                )
                                ws.conditional_formatting.add(f'O{row}', o_rule)
                        elif is_leader_sort or is_leader_stats:
                            # 在团长排序/统计中，检查该玩家的职业是否为素问
                            profession_cell = ws[f'D{row}']  # 职业列
                            if profession_cell.value == '素问':
                                max_o_feather = max(damage_values, key=lambda x: ws[f'O{x[0]}'].value if isinstance(ws[f'O{x[0]}'].value, (int, float)) else 0)[0]
                                max_o_feather_value = ws[f'O{max_o_feather}'].value
                                if max_o_feather_value > 0:
                                    o_rule = DataBarRule(
                                        start_type='num', start_value=0,
                                        end_type='num', end_value=max_o_feather_value,
                                        color='FFC0CB'  # 粉色
                                    )
                                    ws.conditional_formatting.add(f'O{row}', o_rule)
    
    def process(self, output_file=None):
        """主处理函数"""
        if not self.read_csv_data():
            return False
        
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"帮会联赛数据_高级版_{timestamp}.xlsx"
        
        self.create_excel_file(output_file)
        return True

def main_cli(csv_file_path=None):
    """命令行版本的主函数"""
    if csv_file_path is None:
        # 如果没有提供文件路径，使用GUI选择
        return main()
    
    if not os.path.exists(csv_file_path):
        print(f"错误：找不到文件 {csv_file_path}")
        return False
    
    print(f"处理文件：{csv_file_path}")
    
    # 创建处理器并处理数据
    processor = GuildLeagueProcessorAdvanced(csv_file_path)
    
    if processor.process():
        print("数据处理完成！")
        print("生成的文件包含以下工作表：")
        print("1. 关于程序")
        print("2. 本帮团长排序")
        print("3. 本帮职业排序")
        print("4. 敌帮团长排序")
        print("5. 敌帮职业排序")
        print("6. 综合职业排序")
        print("7. 本帮职业统计")
        print("8. 本帮团长统计")
        print("9. 敌帮职业统计")
        print("10. 敌帮团长统计")
        print("11. 帮会对比")
        return True
    else:
        print("数据处理失败！")
        return False

def main():
    """主函数"""
    import tkinter as tk
    from tkinter import filedialog, messagebox
    
    # 创建简单的GUI界面让用户选择文件
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    print("=== 帮会联赛数据处理程序 - 高级版本 ===")
    print("请选择要处理的CSV文件...")
    
    # 打开文件选择对话框
    csv_file = filedialog.askopenfilename(
        title="选择CSV文件",
        filetypes=[
            ("CSV文件", "*.csv"),
            ("所有文件", "*.*")
        ]
    )
    
    if not csv_file:
        print("未选择文件，程序退出。")
        return
    
    print(f"选择的文件：{csv_file}")
    
    if not os.path.exists(csv_file):
        print(f"错误：找不到文件 {csv_file}")
        return
    
    # 创建处理器并处理数据
    processor = GuildLeagueProcessorAdvanced(csv_file)
    
    if processor.process():
        print("数据处理完成！")
        print("生成的文件包含以下工作表：")
        print("1. 关于程序（新增广告页面）")
        print("2. 本帮团长排序")
        print("3. 本帮职业排序")
        print("4. 敌帮团长排序")
        print("5. 敌帮职业排序")
        print("6. 综合职业排序")
        print("7. 本帮职业统计")
        print("8. 本帮团长统计")
        print("9. 敌帮职业统计")
        print("10. 敌帮团长统计")
        print("11. 帮会对比")
        
        # 显示成功消息
        messagebox.showinfo("处理完成", f"数据处理完成！\n输出文件已保存。\n新增了广告页面，包含作者信息和版权声明。")
    else:
        print("数据处理失败！")
        messagebox.showerror("处理失败", "数据处理失败，请检查CSV文件格式。")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        # 命令行模式：python guild_league_processor_advanced.py <csv_file_path>
        csv_file_path = sys.argv[1]
        main_cli(csv_file_path)
    else:
        # GUI模式：python guild_league_processor_advanced.py
        main() 